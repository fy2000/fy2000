# -*- coding: utf-8 -*-
import openpyxl
from modules.opsqlite import SqliteOpt


class CusPaymentInfo:
    def __init__(self, xlsefile: str, dbname: str):
        self._filepath = xlsefile
        self._dbname = dbname
        self._header = {}
        self._datatable = {}
        self._workbook = None
        self._sheet = None

    def read_file(self):
        print(f'start read file {self._filepath}')
        self._workbook = openpyxl.load_workbook(self._filepath)
        self._sheet = self._workbook['客户收款资料']
        self._build_header()
        self._read_file_to_datatable()
        print(f'read file complete! {self._filepath}')

    def _get_tbname(self):
        tbname = 'tb_customer_payment_info'
        return tbname

    def get_data_table(self):
        return self._datatable

    def _build_header(self):
        for i in range(1, self._sheet.max_column + 1):
            headername = self._sheet.cell(row=1, column=i).value
            self._header[headername] = i

    def _read_file_to_datatable(self):
        for row in range(2, self._sheet.max_row + 1):
            # 客户编码 客户名称 简称 单据状态 禁用状态 使用组织 销售组 销售员 结算方式 收款条件 审核人 审核日期 客户分组
            code = self._get_row_cell_value(row, '客户编码')
            fullname = self._get_row_cell_value(row, '客户名称')
            shortname = self._get_row_cell_value(row, '简称')
            status = self._get_row_cell_value(row, '单据状态')
            disable = self._get_row_cell_value(row, '禁用状态')
            organization = self._get_row_cell_value(row, '使用组织')
            salesgroup = self._get_row_cell_value(row, '销售组')
            salesmen = self._get_row_cell_value(row, '销售员')
            settlementModes = self._get_row_cell_value(row, '结算方式')
            paymentterms = self._get_row_cell_value(row, '收款条件')
            reviewer = self._get_row_cell_value(row, '审核人')
            auditdate = self._get_row_cell_value(row, '审核日期')
            cusgrouping = self._get_row_cell_value(row, '客户分组')
            self._datatable[code] = [code, fullname, shortname, status, disable, organization, salesgroup, salesmen,
                                     settlementModes, paymentterms, reviewer, auditdate, cusgrouping]

    def _get_row_cell_value(self, row: str, col: str):
        rvalue = ''
        try:
            rvalue = self._sheet.cell(row=row, column=self._header[col]).value
        except Exception as e:
            rvalue = ''
            print("获取单元格内容失败：" + e.args)
        finally:
            return rvalue

    def save2db(self):
        print('start save data to database...')
        sqlt = SqliteOpt(self._dbname)
        sqlt.save_table_to_db(self._get_tbname(), self.get_data_table())
        sqlt.close()
        print('save data complete!...')

