# -*- coding: utf-8 -*-
import openpyxl
from modules.opsqlite import SqliteOpt


class ReceivableOrderInfo:
    def __init__(self, xlsefile: str, dbname: str):
        self._filepath = xlsefile
        self._dbname = dbname
        self._header = {}
        self._datatable = {}
        self._workbook = None
        self._sheet = None
        self._tbmodel = 'tb_model_receivable_order'
        self._tbprefix = 'tb_receivable_order_'
        self._date = ''

    def read_file(self):
        print(f'start read file {self._filepath}')
        self._workbook = openpyxl.load_workbook(self._filepath)
        self._date = self._filepath.split("_")[1][:4]

        for sheet in self._workbook.sheetnames:
            self._sheet = self._workbook[sheet]
            self._build_header()
            self._read_file_to_datatable()
            print(f'read sheet [{sheet}] complete!')

        print(f'read file complete! {self._filepath}')

    def _get_tbname(self):
        tbname = self._tbprefix + self._date
        return tbname

    def get_data_table(self):
        return self._datatable

    def _build_header(self):
        for i in range(1, self._sheet.max_column + 1):
            headername = self._sheet.cell(row=1, column=i).value
            self._header[headername] = i

    def _read_file_to_datatable(self):
        for row in range(2, self._sheet.max_row + 1):
            # 单据类型 单据编号 业务日期 客户 币别 价税合计 结算组织 销售组织 单据状态 到期日 销售员 物料编码 物料名称 计价单位 计价数量
            # 含税单价 单价 税组合 税率(%) 折扣率(%) 不含税金额 折扣额 税额 价税合计 计价基本数量 不含税金额本位币 表体明细 - 已开票核销金额
            # 销售订单号 收款组织 业务类型 费用承担部门 是否赠品 库存单位 库存数量 库存基本数量 成本金额 已结算金额 客户编码 客户生产单号

            strdocumenttype = self._get_row_cell_value(row, '单据类型')
            if strdocumenttype == '合计':
                continue

            strdocumentcode = self._get_row_cell_value(row, '单据编号')
            strstartdate = self._get_row_cell_value(row, '业务日期')
            strcustomer = self._get_row_cell_value(row, '客户')
            strcurrencytypes = self._get_row_cell_value(row, '币别')
            rtotalpriceincludtax = self._get_row_cell_value(row, '价税合计')
            strsettlementunit = self._get_row_cell_value(row, '结算组织')
            strsalesunit = self._get_row_cell_value(row, '销售组织')
            strdocumentstatus = self._get_row_cell_value(row, '单据状态')
            strlastday = self._get_row_cell_value(row, '到期日')
            strsalesmen = self._get_row_cell_value(row, '销售员')
            strmaterielcode = self._get_row_cell_value(row, '物料编码')
            strmaterielname = self._get_row_cell_value(row, '物料名称')
            strpriceunit = self._get_row_cell_value(row, '计价单位')
            rquantity = self._get_row_cell_value(row, '计价数量')
            runitpriceincludtax = self._get_row_cell_value(row, '含税单价')
            runitprice = self._get_row_cell_value(row, '单价')
            strtaxmix = self._get_row_cell_value(row, '税组合')
            rtaxrate = self._get_row_cell_value(row, '税率(%)')
            rdiscountrate = self._get_row_cell_value(row, '折扣率(%)')
            rtotalprice = self._get_row_cell_value(row, '不含税金额')
            rdiscountprice = self._get_row_cell_value(row, '折扣额')
            rtaxprice = self._get_row_cell_value(row, '税额')
            rtotalpricetax = self._get_row_cell_value(row, '价税合计')
            rbasicquantity = self._get_row_cell_value(row, '计价基本数量')
            rtotalprice_rmb = self._get_row_cell_value(row, '不含税金额本位币')
            rinvoiceamount = self._get_row_cell_value(row, '表体明细 - 已开票核销金额')
            strordernumber = self._get_row_cell_value(row, '销售订单号')
            strpayee = self._get_row_cell_value(row, '收款组织')
            strbusinessstype = self._get_row_cell_value(row, '业务类型')
            strcostbearingDepartment = self._get_row_cell_value(row, '费用承担部门')
            strgift = self._get_row_cell_value(row, '是否赠品')
            strinventoryunit = self._get_row_cell_value(row, '库存单位')
            rinventoryquantity = self._get_row_cell_value(row, '库存数量')
            rbaseinventoryquantity = self._get_row_cell_value(row, '库存基本数量')
            rcost = self._get_row_cell_value(row, '成本金额')
            rSettledamount = self._get_row_cell_value(row, '已结算金额')
            strcustomercode = self._get_row_cell_value(row, '客户编码')
            strproductioncode = self._get_row_cell_value(row, '客户生产单号')

            self._datatable[len(self._datatable)] = [None, strdocumenttype, strdocumentcode, strstartdate, strcustomer,
                                                     strcurrencytypes, rtotalpriceincludtax, strsettlementunit,
                                                     strsalesunit, strdocumentstatus, strlastday, strsalesmen,
                                                     strmaterielcode, strmaterielname, strpriceunit, rquantity,
                                                     runitpriceincludtax, runitprice, strtaxmix, rtaxrate,
                                                     rdiscountrate, rtotalprice, rdiscountprice, rtaxprice,
                                                     rtotalpricetax, rbasicquantity, rtotalprice_rmb, rinvoiceamount,
                                                     strordernumber, strpayee, strbusinessstype,
                                                     strcostbearingDepartment, strgift, strinventoryunit,
                                                     rinventoryquantity, rbaseinventoryquantity, rcost, rSettledamount,
                                                     strcustomercode, strproductioncode]

    def _get_row_cell_value(self, row: str, col: str):
        rvalue = ''
        try:
            rvalue = self._sheet.cell(row=row, column=self._header[col]).value
        except Exception as e:
            rvalue = ''
            print("获取单元格内容失败：" + e.args)
        finally:
            return rvalue

    def save2db(self):
        print('start save data to database...')
        sqlt = SqliteOpt(self._dbname)
        sqlt.create_table_by_tbmodel(self._tbmodel, self._get_tbname())
        sqlt.save_table_to_db(self._get_tbname(), self.get_data_table())
        sqlt.close()
        print('save data complete!...')
