# -*- coding: utf-8 -*-
import openpyxl
from modules.opsqlite import SqliteOpt


class AnnualSalesInfo:
    def __init__(self, xlsefile: str, dbname: str):
        self._filepath = xlsefile
        self._dbname = dbname
        self._header = {}
        self._datatable = {}
        self._workbook = None
        self._sheet = None
        self._tbmodel = 'tb_model_total_sales_order'
        self._tbprefix = 'tb_total_sales_order_'
        self._date = ''

    def read_file(self):
        print(f'start read file {self._filepath}')
        self._workbook = openpyxl.load_workbook(self._filepath)
        self._date = self._workbook.active.title[:4]

        self._sheet = self._workbook[self._workbook.active.title]
        self._build_header()
        self._read_file_to_datatable()
        print(f'read file complete! {self._filepath}')

    def _get_tbname(self):
        tbname = self._tbprefix + self._date
        return tbname

    def get_data_table(self):
        return self._datatable

    def _build_header(self):
        for i in range(1, self._sheet.max_column + 1):
            headername = self._sheet.cell(row=1, column=i).value
            self._header[headername] = i

    def _read_file_to_datatable(self):
        for row in range(2, self._sheet.max_row + 1):
            # 日期 单据类型 单据编号 单据状态 客户 销售部门 销售员 关闭状态 物料编码 物料名称 规格型号 销售单位 销售数量 单价 含税单价
            # 价税合计 要货日期 累计出库数量（销售基本） 累计退货数量（销售基本） 金额（本位币） 价税合计（本位币） 剩余未出数量（销售基本）
            # 业务关闭 循环备品 客户生产单号

            strdate = self._get_row_cell_value(row, '日期')
            if len(strdate.split("/")) != 3:
                continue
            strdocumenttype = self._get_row_cell_value(row, '单据类型')
            strdocumentcode = self._get_row_cell_value(row, '单据编号')
            strdocumentstatus = self._get_row_cell_value(row, '单据状态')
            strcustomer = self._get_row_cell_value(row, '客户')
            strsalesdepartment = self._get_row_cell_value(row, '销售部门')
            strsalesmen = self._get_row_cell_value(row, '销售员')
            strclosedstatus = self._get_row_cell_value(row, '关闭状态')
            strmaterielcode = self._get_row_cell_value(row, '物料编码')
            strmaterielname = self._get_row_cell_value(row, '物料名称')
            strspecifications = self._get_row_cell_value(row, '规格型号')
            strsalesunit = self._get_row_cell_value(row, '销售单位')
            isalesnums = self._get_row_cell_value(row, '销售数量')
            runitprice = self._get_row_cell_value(row, '单价')
            runitpriceincludtax = self._get_row_cell_value(row, '含税单价')
            rtotalpriceincludtax = self._get_row_cell_value(row, '价税合计')
            strrequiredate = self._get_row_cell_value(row, '要货日期')
            rcumulativedelivery = self._get_row_cell_value(row, '累计出库数量（销售基本）')
            rcumulativereturn = self._get_row_cell_value(row, '累计退货数量（销售基本）')
            rtotalprice_rmb = self._get_row_cell_value(row, '金额（本位币）')
            rtotalpriceincludtax_rmb = self._get_row_cell_value(row, '价税合计（本位币）')
            irestquantity = self._get_row_cell_value(row, '剩余未出数量（销售基本）')
            strbusinessstatus = self._get_row_cell_value(row, '业务关闭')
            strcyclespare = self._get_row_cell_value(row, '循环备品')
            strproductioncode = self._get_row_cell_value(row, '客户生产单号')
            self._datatable[len(self._datatable)] = [None, strdate, strdocumenttype, strdocumentcode, strdocumentstatus,
                                                     strcustomer, strsalesdepartment, strsalesmen, strclosedstatus,
                                                     strmaterielcode, strmaterielname, strspecifications, strsalesunit,
                                                     isalesnums, runitprice, runitpriceincludtax, rtotalpriceincludtax,
                                                     strrequiredate, rcumulativedelivery, rcumulativereturn,
                                                     rtotalprice_rmb, rtotalpriceincludtax_rmb, irestquantity,
                                                     strbusinessstatus, strcyclespare, strproductioncode]

    def _get_row_cell_value(self, row: str, col: str):
        rvalue = ''
        try:
            rvalue = self._sheet.cell(row=row, column=self._header[col]).value
        except Exception as e:
            rvalue = ''
            print("获取单元格内容失败：" + e.args)
        finally:
            return rvalue

    def save2db(self):
        print('start save data to database...')
        sqlt = SqliteOpt(self._dbname)
        sqlt.create_table_by_tbmodel(self._tbmodel, self._get_tbname())
        sqlt.save_table_to_db(self._get_tbname(), self.get_data_table())
        sqlt.close()
        print('save data complete!...')
