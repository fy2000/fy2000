# -*- coding:utf-8 -*-

import os
import openpyxl


class ExcelOpt:
    def __init__(self, file_name):
        self.__filename = None
        self.__wb = None
        self.__ws = None
        self._header = {}

        if os.path.exists(file_name):
            self.__filename = file_name
            self.__wb = openpyxl.load_workbook(self.__filename)
        else:
            print('Excel file is not exist! please check ...')

    def _build_header(self):
        if self.__ws:
            for icol in range(1, self.__ws.max_column + 1):
                header_name = self.__ws.cell(row=1, column=icol).value
                self._header[header_name] = icol

    def _get_row_cell_value(self, row: str, col: str):
        rvalue = ''
        try:
            rvalue = self.__ws.cell(row=row, column=self._header[col]).value
        except Exception as e:
            rvalue = ''
            print("获取单元格内容失败：" + e.args)
        finally:
            return rvalue

    def read_excel(self, sheetname: str):
        if self.__wb is None:
            return

        if (not sheetname) and (sheetname not in self.__wb.sheetnames):
            print('Given sheet name is wrong! please check ...')
        else:
            self.__ws = self.__wb[sheetname]
            self._build_header()

    def get_value(self, row_num, col_name: str):
        return self._get_row_cell_value(row_num, col_name)

    def max_row(self):
        if self.__ws:
            return self.__ws.max_row
        else:
            return 0

    def max_column(self):
        if self.__ws:
            return self.__ws.max_column
        else:
            return 0


if __name__ == '__main__':
    my_excel = ExcelOpt(r"C:\Users\mt\Desktop\统计汇总\LTE数据.xlsx")
    my_excel.read_excel("Sheet1")
    for i in range(2, my_excel.max_row() + 1):
        print(my_excel.get_value(i, '文件名'))
