# -*- coding : utf-8 -*-

from gevent import monkey

monkey.patch_all()

import os
import requests
from gevent import queue
from bs4 import BeautifulSoup
import gevent
import openpyxl


def get_header():
    headers = {
        'User-Agent': 'Mozilla/5.0(Windows NT 10.0; WOW64) AppleWebKit/537.36(KHTML, like Gecko) '
                      'Chrome/84.0.4147.105 Safari/537.36'
    }

    return headers


def get_food_group_url():
    group_list = []
    for group_num in range(1, 4):
        for page_num in range(1, 4):
            url_food = 'http://www.boohee.com/food/group/{0}?page={1}'.format(group_num, page_num)
            group_list.append(url_food)
    return group_list


def get_dish_group_url():
    group_list = []
    for page_num in range(1, 4):
        url_dish = 'http://www.boohee.com/food/view_menu?page={0}'.format(page_num)
        group_list.append(url_dish)
    return group_list


def execute_fun(wq: gevent._gevent_cqueue.Queue):
    ret_info = list()
    while not wq.empty():
        f_url = wq.get_nowait()
        response = requests.get(f_url, headers=get_header())
        if response.status_code == 200:
            bs = BeautifulSoup(response.text, 'html.parser')
            bs_div = bs.find('div', class_='widget-food-list')
            group_name = bs_div.find('h3').text.strip()
            food_list_info = bs_div.find('ul', class_='food-list').find_all('li', class_='item')
            for f_item in food_list_info:
                f_div = f_item.find('div', class_='text-box')
                f_name = f_div.h4.a.text
                f_power = f_div.p.text
                f_link = 'http://www.boohee.com' + f_div.h4.a['href']
                ret_info.append([group_name, f_name, f_power, f_link])
    return ret_info


def save_to_excel(filename: str, data_l: list):
    print('开始存储数据到 {0} 文件...'.format(filename))

    # 再次存储之前，需要先删除之前文件
    if os.path.exists(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])

    for dt in data_l:
        group_name = dt[0]
        if group_name in wb.sheetnames:
            sheet = wb[group_name]
            sheet.append(dt)
        else:
            sheet = wb.create_sheet(group_name)
            sheet.append(['所属类型', '食物名称', '热量', '详情链接'])
            sheet.append(dt)

    wb.save(filename)
    wb.close()
    print('数据存储完毕...')


if __name__ == "__main__":
    foods_list = get_food_group_url()
    dish_list = get_dish_group_url()
    url_list = foods_list + dish_list

    # 创建队列对象，并添加请求的网址到队列中
    work_q = queue.Queue()
    for item in url_list:
        work_q.put_nowait(item)

    # 建立三个爬虫任务
    task_list = list()
    for i in range(1, 4):
        task = gevent.spawn(execute_fun, work_q)
        task_list.append(task)

    print('开始爬取热量数据...')
    gevent.joinall(task_list)
    print('数据爬取完毕...')

    # 整合各个爬虫任务爬取的数据
    data_list = list()
    for g_task in task_list:
        data_list += g_task.value

    print('总共爬取数据 {0} 条...'.format(len(data_list)))

    # 存储数据...
    save_to_excel('食物热量表.xlsx', data_list)

    print('程序执行结束...')
